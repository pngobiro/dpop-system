import os
import openpyxl
from django.conf import settings
from openpyxl.utils import get_column_letter
from .models import DcrtData # Keep original model import
import logging

logger = logging.getLogger(__name__)

# Helper function to safely get value from row tuple
def get_value_from_row(row, index, expected_type=None, default=None):
    """Safely gets a value from a row tuple, attempts type conversion, and handles errors."""
    raw_value = None
    try:
        if index < len(row):
            raw_value = row[index]
            if raw_value is None:
                return default

            if expected_type == int:
                # Attempt to convert to int, handle potential errors
                try:
                    # Handle cases like '1.0' which are valid floats but should be int
                    if isinstance(raw_value, float) and raw_value.is_integer():
                         return int(raw_value)
                    return int(raw_value)
                except (ValueError, TypeError):
                    logger.warning(f"Could not convert '{raw_value}' (type: {type(raw_value)}) to int for index {index}. Returning default.")
                    return default # Return default (None) if conversion fails
            elif expected_type == str:
                 return str(raw_value) # Ensure string type if needed
            # Add other type checks (float, date) if necessary
            else:
                return raw_value # Return raw value if no specific type needed
        else:
            # Index out of range
            return default
    except IndexError:
        # This case should be less likely with the len check, but keep for safety
        return default
    except Exception as e:
        logger.error(f"Unexpected error in get_value_from_row for index {index}, value '{raw_value}': {e}")
        return default
def handle_uploaded_file(uploaded_file, unit_rank, fy, fq, unit, division, month):
    """
    Handles uploaded Excel file, saves it to a structured path, and processes its data.
    Path: MEDIA_ROOT/TEMPLATE <rank_id>/<unit_name>/<year_name>/<month_number>/<unit_code>.xlsx
    """
def get_dcrt_filepath(unit_rank, fy, month, unit, base_dir=os.path.join(settings.BASE_DIR, 'DCRT'), extension=".xlsx"):
    """
    Constructs the structured filepath for a DCRT Excel file.
    Pattern: DCRT/TEMPLATE {rank_id}/{unit name}/{year}/{month}/{unique_code}.xlsx
    Example: DCRT/TEMPLATE 4/Kisumu ELRC/2024/01/KSMERC.xlsx
    """
    try:
        import glob

        # Map month names to their numbers in file system
        month_numbers = {
            'July': '07',
            'August': '08',
            'September': '09',
            'October': '10',
            'November': '11',
            'December': '12',
            'January': '01',
            'February': '02',
            'March': '03',
            'April': '04',
            'May': '05',
            'June': '06'
        }
        month_str = month_numbers[month.name]
        
        # For Q1 months (July-September), use the first year of financial year
        # For other months, use second year
        if month.name in ['July', 'August', 'September']:
            year = fy.name.split('/')[0]  # Use first year (e.g., "2024" from "2024/2025")
        else:
            year = fy.name.split('/')[1]  # Use second year (e.g., "2025" from "2024/2025")
            
        logger.info(f"Processing month: {month.name} ({month_str}), year: {year}")
        year = fy.name.split('/')[0]
        
        # First look for files in this unit's directory
        unit_pattern = os.path.join(
            base_dir,
            f"TEMPLATE {unit_rank.id}",
            f"*{unit.name}*",  # Allow partial matches of unit name
            year,
            month_str,
            "*.xlsx"
        )
        logger.info(f"Checking unit pattern: {unit_pattern}")
        
        # Look for existing files in unit's directory
        unit_files = glob.glob(unit_pattern)
        if unit_files:
            logger.info(f"Found file in unit directory: {unit_files[0]}")
            return unit_files[0]
        
        # Look for template files in other court directories
        template_pattern = os.path.join(
            base_dir,
            f"TEMPLATE {unit_rank.id}",
            "*",  # Any court directory
            year,
            month_str,
            "*.xlsx"
        )
        logger.info(f"Checking template pattern: {template_pattern}")
        
        # Look for template files
        template_files = glob.glob(template_pattern)
        if template_files:
            logger.info("Available template files:")
            for template in template_files:
                court_dir = os.path.basename(os.path.dirname(os.path.dirname(template)))
                logger.info(f"- {court_dir}: {os.path.basename(template)}")
            selected_template = template_files[0]
            logger.info(f"Using template file: {selected_template}")
            return selected_template
        
        # If no existing files found, construct expected path for new file
        expected_path = os.path.join(
            base_dir,
            f"TEMPLATE {unit_rank.id}",
            unit.name,
            year,
            month_str,
            f"{unit.unique_code}{extension}"
        )
        logger.info(f"No existing files found. Using new path: {expected_path}")
        return expected_path

    except AttributeError as e:
        logger.error(f"Invalid attribute access while constructing path: {e}")
        return None
    except Exception as e:
        logger.error(f"Error generating DCRT filepath: {e}")
        return None
        unit_name = unit.name
        
        # Construct the relative path components
        relative_path = os.path.join(
            f"TEMPLATE {unit_rank.id}",  # e.g., "TEMPLATE 4"
            unit_name,                    # e.g., "Kisumu ELRC"
            year,                         # e.g., "2024"
            f"{month.month_number:02d}"   # e.g., "01"
        )
        
        # Construct the filename using unit unique code
        filename = f"{unit.unique_code}{extension}"  # e.g., "KSMERC.xlsx"
        
        # Return the full path
        return os.path.join(base_dir, relative_path, filename)
        
    except Exception as e:
        logger.error(f"Error generating DCRT filepath for unit {unit.id}, month {month.id}: {e}")
        return None

def handle_uploaded_file(uploaded_file, unit_rank, fy, fq, unit, division, month):
    """
    Handles uploaded Excel file, saves it to a structured path, and processes its data.
    Path: MEDIA_ROOT/TEMPLATE <rank_id>/<unit_name>/<year_name>/<month_number>/<unit_code>.xlsx
    """
    try:
        # Use the helper function to get the destination path
        destination_path = get_dcrt_filepath(unit_rank, fy, month, unit, extension=os.path.splitext(uploaded_file.name)[1])
        
        if not destination_path:
            logger.error("Could not determine destination path for upload.")
            return None # Indicate save failure

        # Ensure the directory exists
        full_dir_path = os.path.dirname(destination_path)
        os.makedirs(full_dir_path, exist_ok=True)
        
        logger.info(f"Saving uploaded file to: {destination_path}")

        # Save the file
        with open(destination_path, 'wb+') as destination_file:
            for chunk in uploaded_file.chunks():
                destination_file.write(chunk)
    except Exception as e:
        logger.error(f"Error saving uploaded file: {e}")
        # Decide how to handle save error (e.g., raise, return False)
        return None # Indicate save failure

    # --- Existing Excel processing logic ---
    try:
        # Load the saved Excel file
        workbook = openpyxl.load_workbook(destination_path)
        worksheet = workbook.active
        logger.info(f"Processing data from: {destination_path}")

        # Start reading data from row 6, column B
        row_count = 0
        for row_index, row in enumerate(worksheet.iter_rows(min_row=6, min_col=2, values_only=True), start=6):
            if all([cell is None for cell in row]):
                logger.info(f"Stopping processing at empty row {row_index}")
                break
            
            # Create a DcrtData object using the helper function for safety
            # Build data dictionary with type conversion attempts
            data_dict = {
                'today_date_day': get_value_from_row(row, 0, expected_type=int),
                'today_date_month': get_value_from_row(row, 1),
                'today_date_year': get_value_from_row(row, 2),
                'case_number_code': get_value_from_row(row, 3),
                'case_number_number': get_value_from_row(row, 4, expected_type=int),
                'case_number_day': get_value_from_row(row, 5, expected_type=int),
                'case_number_month': get_value_from_row(row, 6),
                'case_number_year': get_value_from_row(row, 7, expected_type=int),
                'appeal_number_court_name': get_value_from_row(row, 8),
                'appeal_number_code': get_value_from_row(row, 9),
                'appeal_number_number': get_value_from_row(row, 10, expected_type=int),
                'appeal_number_year': get_value_from_row(row, 11, expected_type=int),
                'specific_case_type': get_value_from_row(row, 12),
                'judicial_officer_1': get_value_from_row(row, 13),
                'judicial_officer_2': get_value_from_row(row, 14),
                'judicial_officer_3': get_value_from_row(row, 15),
                'judicial_officer_4': get_value_from_row(row, 16),
                'judicial_officer_5': get_value_from_row(row, 17),
                'judicial_officer_6': get_value_from_row(row, 18),
                'judicial_officer_7': get_value_from_row(row, 19),
                'judicial_officer_8': get_value_from_row(row, 20),
                'case_coming_for': get_value_from_row(row, 21),
                'case_outcome': get_value_from_row(row, 22),
                'adjournment_reason': get_value_from_row(row, 23),
                'date_of_next_activity_day': get_value_from_row(row, 24, expected_type=int),
                'date_of_next_activity_month': get_value_from_row(row, 25),
                'date_of_next_activity_year': get_value_from_row(row, 26, expected_type=int),
                'no_of_plaintiffs_or_appellants_male': get_value_from_row(row, 27, expected_type=int),
                'no_of_plaintiffs_or_appellants_female': get_value_from_row(row, 28, expected_type=int),
                'no_of_plaintiffs_or_appellants_organization': get_value_from_row(row, 29, expected_type=int),
                'no_of_defendants_accused_male': get_value_from_row(row, 30, expected_type=int),
                'no_of_defendants_accused_female': get_value_from_row(row, 31, expected_type=int),
                'no_of_defendants_accused_organization': get_value_from_row(row, 32, expected_type=int),
                'parties_have_legal_representation': get_value_from_row(row, 33),
                'no_of_witnesses_in_court_d': get_value_from_row(row, 34, expected_type=int),
                'no_of_witnesses_in_court_w': get_value_from_row(row, 35, expected_type=int),
                'no_of_accused_remanded': get_value_from_row(row, 36, expected_type=int),
                'last_date_of_submission_of_case_file_day': get_value_from_row(row, 37),
                'last_date_of_submission_of_case_file_month': get_value_from_row(row, 38),
                'last_date_of_submission_of_case_file_year': get_value_from_row(row, 39),
                'remarks': get_value_from_row(row, 40),
            }

            # Create and save DcrtData object
            try:
                DcrtData.objects.create(
                    unit=unit,
                    financial_year=fy,
                    financial_quarter=fq,
                    month=month,
                    division=division,
                    **data_dict # Unpack the cleaned dictionary
                )
                row_count += 1
            except Exception as e:
                # Log the specific data that caused the save error
                logger.error(f"Error saving row {row_index} data: {data_dict}. Error: {e}")
                # Decide whether to continue processing other rows or stop
                # continue # or break or raise e
        logger.info(f"Successfully processed {row_count} rows from Excel file.")
    except Exception as e:
        logger.error(f"Error processing Excel file {destination_path}: {e}")
        return None # Indicate processing failure

    # Removed duplicated code block that was causing NameError
